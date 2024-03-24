import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def no_fill(active_sheet_name):  #Takes a specific sheet as an argument and sets the positions table to no fill
    ws = wb[active_sheet_name]
    blank_fill = PatternFill(fill_type="none")

    for col_index in range(9, 17):
        for row in range(3, 166):
            cell = ws.cell(row=row, column=col_index)
            cell.fill = blank_fill

no_fill("Atlanta")

wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')



import openpyxl
from openpyxl.styles import PatternFill

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def start_count_and_highlight(name, sec_name, active_sheet_name, column_index, target_col):
    #Yellow fill for players not designated as a usual starter in position table
    #Also tallies a number of starts for the players entered as arguments
    count = 0
    ws = wb[active_sheet_name]

    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=3, max_row=166, min_col=column_index, max_col=column_index):
        cell = row[0]  # Access the cell from the row
        cell_value = cell.value

        if cell_value != name and cell_value != sec_name:
            count += 1
            cell.fill = yellow_fill

    ws.cell(row=168, column=target_col, value=count)

    return count

# Call the function with the desired column index and target row
print(start_count_and_highlight('Stearns', '', active_sheet_name='New York', column_index=9, target_col=9))
print(start_count_and_highlight('Montanez', '', active_sheet_name='New York', column_index=10, target_col=10))
print(start_count_and_highlight('Flynn', '', active_sheet_name='New York', column_index=11, target_col=11))
print(start_count_and_highlight('Hebner', '', active_sheet_name='New York', column_index=12, target_col=12))
print(start_count_and_highlight('Taveras', '', active_sheet_name='New York', column_index=13, target_col=13))
print(start_count_and_highlight('Henderson', 'Youngblood', active_sheet_name='New York', column_index=14, target_col=14))
print(start_count_and_highlight('Mazzilli', 'Youngblood', active_sheet_name='New York', column_index=15, target_col=15))
print(start_count_and_highlight('Youngblood', '', active_sheet_name='New York', column_index=16, target_col=16))

# Save the workbook
#wb.save(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')



import openpyxl

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def player_list(sheet_name, column, position): #Returns a list with the name of each player who started a game at a given position (Console)
    active_sheet = wb[sheet_name]
    plyr_list = []

    for row in active_sheet.iter_rows(min_row=3, max_row=162, min_col=column, max_col=column):
        cell = row[0]
        cell_value = cell.value

        if cell_value not in plyr_list:
            plyr_list.append(cell_value)

    result = f"{position}: {plyr_list}"

    return result

result_9 = player_list(sheet_name="ST Louis", column=9, position="Catcher")
result_10 = player_list(sheet_name="ST Louis", column=10, position="First Base")
result_11 = player_list(sheet_name="ST Louis", column=11, position="Second Base")
result_12 = player_list(sheet_name="ST Louis", column=12, position="Third Base")
result_13 = player_list(sheet_name="ST Louis", column=13, position="Shortstop")
result_14 = player_list(sheet_name="ST Louis", column=14, position="Left Field")
result_15 = player_list(sheet_name="ST Louis", column=15, position="Centerfiedl")
result_16 = player_list(sheet_name="ST Louis", column=16, position="Right Field")
result_17 = player_list(sheet_name="ST Louis", column=17, position="Pitcher")

print(result_9)
print(result_10)
print(result_11)
print(result_12)
print(result_13)
print(result_14)
print(result_15)
print(result_16)
print(result_17)  #Column 9 is for catchers

#wb.save(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

import openpyxl
from openpyxl.styles import PatternFill

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def pitcher_highlight(sheet_name, column, name): # Removes fill in pitcher column. Then highlights all starts for pitcher (name)
    active_sheet = wb[sheet_name]

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    no_fill = PatternFill(fill_type="none")

    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=column, max_col=column):
        cell = row[0]
        cell.fill = no_fill
    
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=column, max_col=column):
        cell = row[0]
        cell_value = cell.value

        if cell_value == name:
            cell.fill = green_fill

        if cell_value != name:
            cell.fill = no_fill

pitcher_highlight(sheet_name="Cincinnati", column=17, name="Norman")
pitcher_highlight(sheet_name="Philadelphia", column=17, name="Lerch")
pitcher_highlight(sheet_name="ST Louis", column=17, name="Denny")
pitcher_highlight(sheet_name="San Francisco", column=17, name="Knepper")
pitcher_highlight(sheet_name="Atlanta", column=17, name="Brizzolara")
pitcher_highlight(sheet_name="New York", column=17, name="Falcone")
pitcher_highlight(sheet_name="Chicago", column=17, name="Krukow")
pitcher_highlight(sheet_name="San Diego", column=17, name="Owchinko")
pitcher_highlight(sheet_name="Houston", column=17, name="Niekro")
pitcher_highlight(sheet_name="Montreal", column=17, name="Schatzeder")
pitcher_highlight(sheet_name="Pittsburgh", column=17, name="Whitson")
pitcher_highlight(sheet_name="Los Angeles", column=17, name="Reuss")

#wb.save(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')



import openpyxl

# Assuming 'wb' is a global variable pointing to the workbook
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def start_count(sheet_name, pitcher, date):  # Returns start number and total starts for pitcher. Then places values in More_info sheet
    current_date_row = 0
    active_sheet = wb[sheet_name]

    # Find the row corresponding to the current date
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=8, max_col=8):
        cell = row[0]
        if date in str(cell.value) and current_date_row == 0:
            current_date_row = cell.row

    start_count = 0
    total_starts = 0

    # Count starts and total starts
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=17, max_col=17):
        cell = row[0]
        cell_value = cell.value

        if pitcher in str(cell_value):
            total_starts += 1

        if cell.row <= current_date_row and pitcher in str(cell_value):
            start_count += 1

    return start_count, total_starts

# Example usage
data_sets = [
     {"sheet_name": "Montreal", "date": "6/18", "pitcher": "Rogers"},
     {"sheet_name": "ST Louis", "date": "6/18", "pitcher": "Forsch"},
     {"sheet_name": "Philadelphia", "date": "6/18", "pitcher": "Espinosa"},
     {"sheet_name": "Pittsburgh", "date": "6/18", "pitcher": ""},
     {"sheet_name": "New York", "date": "6/18", "pitcher": "Ellis"},
     {"sheet_name": "Atlanta", "date": "6/18", "pitcher": "Matula"},
     {"sheet_name": "Cincinnati", "date": "6/18", "pitcher": "Bonham"},
     {"sheet_name": "Houston", "date": "6/18", "pitcher": "Williams"},
     {"sheet_name": "San Francisco", "date": "6/18", "pitcher": ""},
     {"sheet_name": "San Diego", "date": "6/18", "pitcher": "Jones"},
     {"sheet_name": "Los Angeles", "date": "6/18", "pitcher": "Hooton"},
     {"sheet_name": "Chicago", "date": "6/18", "pitcher": "Lamp"},
]

active_sheet = wb['More_Info']

row_counter = 1

for data in data_sets:
    pitcher = data["pitcher"]
    date = data["date"]
    team = data["sheet_name"]
    start_count_value, total_starts_value = start_count(team, pitcher, date)

    active_sheet.cell(row=row_counter, column=1).value = date
    active_sheet.cell(row=row_counter, column=2).value = team
    active_sheet.cell(row=row_counter, column=3).value = pitcher
    active_sheet.cell(row=row_counter, column=4).value = start_count_value
    active_sheet.cell(row=row_counter, column=5).value = total_starts_value
    
    if round(float(active_sheet.cell(row=row_counter, column=4).value)) == 0:
        active_sheet.cell(row=row_counter, column=4).value = "No Game"
    if round(float(active_sheet.cell(row=row_counter, column=5).value)) == 163:
        active_sheet.cell(row=row_counter, column=5).value = "Scheduled"

    
    row_counter += 1

# Save the workbook to apply changes
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

# Close the workbook
wb.close() 


import openpyxl
from openpyxl.styles import PatternFill

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_AL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def find_player(sheet_name, date, column): # Returns player name for arguments of date and column (Console).  Not sure how I use this.
    current_date_row = 0
    active_sheet = wb[sheet_name]

    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=8, max_col=8):
        cell = row[0]
        if date in str(cell.value) and current_date_row == 0:
            current_date_row = cell.row
            value_in_column = active_sheet.cell(row=current_date_row, column=column).value
            return value_in_column
        
print(find_player("Chicago", "8/21", 12))
        
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')


# Assuming 'wb' is a global variable pointing to the workbook
# Assuming 'wb' is a global variable pointing to the workbook
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def start_count(sheet_name, pitcher, date):  # Returns start number and total starts for pitcher. Then places values in More_info sheet
    current_date_row = 0
    active_sheet = wb[sheet_name]

    # Find the row corresponding to the current date
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=8, max_col=8):
        cell = row[0]
        if date in str(cell.value) and current_date_row == 0:
            current_date_row = cell.row

    start_count = 0
    total_starts = 0

    # Count starts and total starts
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=17, max_col=17):
        cell = row[0]
        cell_value = cell.value

        if pitcher in str(cell_value):
            total_starts += 1

        if cell.row <= current_date_row and pitcher in str(cell_value):
            start_count += 1

    return start_count, total_starts

# Example usage
data_sets = [
    # Your data sets here...
]

active_sheet = wb['More_Info']

row_counter = 1

for data in data_sets:
    pitcher = data["pitcher"]
    date = data["date"]
    team = data["sheet_name"]
    start_count_value, total_starts_value = start_count(team, pitcher, date)

    active_sheet.cell(row=row_counter, column=1).value = date
    active_sheet.cell(row=row_counter, column=2).value = team
    active_sheet.cell(row=row_counter, column=3).value = pitcher
    try:
        # Convert to float first to handle decimal places
        cell_value_column4 = int(float(active_sheet.cell(row=row_counter, column=4).value))
        if cell_value_column4 != 0:
            active_sheet.cell(row=row_counter, column=4).value = start_count_value
        else:
            active_sheet.cell(row=row_counter, column=4).value = "No Game"
    except ValueError:
        # Handle the case where conversion is not possible
        active_sheet.cell(row=row_counter, column=4).value = "Invalid"
    
    try:
        # Convert to float first to handle decimal places
        cell_value_column5 = int(float(active_sheet.cell(row=row_counter, column=5).value))
        if cell_value_column5 < 50:
            active_sheet.cell(row=row_counter, column=5).value = total_starts_value
        else:
            active_sheet.cell(row=row_counter, column=5).value = "Scheduled"
    except ValueError:
        # Handle the case where conversion is not possible
        active_sheet.cell(row=row_counter, column=5).value = "Invalid"
    
    row_counter += 1

# Save the workbook to apply changes
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

# Close the workbook
wb.close()

import openpyxl
from openpyxl.styles import PatternFill
# Assuming 'wb' is a global variable pointing to the workbook
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def app_count(sheet_name, plr_name):
    active_sheet = wb[sheet_name]
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

    start_row = 3
    end_row = 173
    column = 'G'

     # Iterate through the specified column and highlight rows with player's name
    for row in range(start_row, end_row + 1):
        cell_value = active_sheet[f'{column}{row}'].value
        if plr_name in str(cell_value):
            # Apply yellow fill to the entire row
            for cell in active_sheet[row]:
                active_sheet[f'{column}{row}'].fill = yellow_fill

app_count('Chicago', 'Wortham')

wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

# Close the workbook
wb.close()
