import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def no_fill(active_sheet_name):
    ws = wb[active_sheet_name]
    blank_fill = PatternFill(fill_type="none")

    for col_index in range(9, 17):
        for row in range(3, 166):
            cell = ws.cell(row=row, column=col_index)
            cell.fill = blank_fill

no_fill("Atlanta")

wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

import openpyxl
from openpyxl.styles import PatternFill

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def start_count_and_highlight(name, sec_name, active_sheet_name, column_index, target_col):
    count = 0
    ws = wb[active_sheet_name]

    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

    for row in ws.iter_rows(min_row=3, max_row=166, min_col=column_index, max_col=column_index):
        cell = row[0]  # Access the cell from the row
        cell_value = cell.value

        if cell_value != name and cell_value != sec_name:
            count += 1
            cell.fill = yellow_fill

    ws.cell(row=168, column=target_col, value=count)

    return count

# Call the function with the desired column index and target row
print(start_count_and_highlight('Stearns', '', active_sheet_name='New York', column_index=9, target_col=9))
print(start_count_and_highlight('Montanez', '', active_sheet_name='New York', column_index=10, target_col=10))
print(start_count_and_highlight('Flynn', '', active_sheet_name='New York', column_index=11, target_col=11))
print(start_count_and_highlight('Hebner', '', active_sheet_name='New York', column_index=12, target_col=12))
print(start_count_and_highlight('Taveras', '', active_sheet_name='New York', column_index=13, target_col=13))
print(start_count_and_highlight('Henderson', 'Youngblood', active_sheet_name='New York', column_index=14, target_col=14))
print(start_count_and_highlight('Mazzilli', 'Youngblood', active_sheet_name='New York', column_index=15, target_col=15))
print(start_count_and_highlight('Youngblood', '', active_sheet_name='New York', column_index=16, target_col=16))

# Save the workbook
#wb.save(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

import openpyxl

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def player_list(sheet_name, column):
    active_sheet = wb[sheet_name]
    plyr_list = []

    for row in active_sheet.iter_rows(min_row=3, max_row=162, min_col=column, max_col=column):
        cell = row[0]
        cell_value = cell.value

        if cell_value not in plyr_list:
            plyr_list.append(cell_value)

    return plyr_list

print(player_list(sheet_name="Atlanta", column=14))  #Column 13 is for catchers

#wb.save(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

import openpyxl
from openpyxl.styles import PatternFill

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def pitcher_highlight(sheet_name, column, name):
    active_sheet = wb[sheet_name]

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    no_fill = PatternFill(fill_type="none")

    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=column, max_col=column):
        cell = row[0]
        cell.fill = no_fill
    
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=column, max_col=column):
        cell = row[0]
        cell_value = cell.value

        if cell_value == name:
            cell.fill = green_fill

        if cell_value != name:
            cell.fill = no_fill

pitcher_highlight(sheet_name="Cincinnati", column=17, name="Moskau")
pitcher_highlight(sheet_name="Philadelphia", column=17, name="Carlton")
pitcher_highlight(sheet_name="ST Louis", column=17, name="Martinez")
pitcher_highlight(sheet_name="San Francisco", column=17, name="Curtis")
pitcher_highlight(sheet_name="Atlanta", column=17, name="Mahler")
pitcher_highlight(sheet_name="New York", column=17, name="Kobel")
pitcher_highlight(sheet_name="Chicago", column=17, name="Holtzman")
pitcher_highlight(sheet_name="San Diego", column=17, name="Perry")
pitcher_highlight(sheet_name="Houston", column=17, name="Richard")
pitcher_highlight(sheet_name="Montreal", column=17, name="Grimsley")
pitcher_highlight(sheet_name="Pittsburgh", column=17, name="Robinson")
pitcher_highlight(sheet_name="Los Angeles", column=17, name="Welch")

#wb.save(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_NL_Schedule-By_Team.xlsx')
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

import openpyxl
from openpyxl.styles import PatternFill

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_AL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_AL_Schedule-By_Team.xlsx')

def pitcher_highlight(sheet_name, column, name):
    active_sheet = wb[sheet_name]
    count = 0

    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    no_fill = PatternFill(fill_type="none")

    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=column, max_col=column):
        cell = row[0]
        cell.fill = no_fill
    
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=column, max_col=column):
        cell = row[0]
        cell_value = cell.value

        if cell_value == name:
            cell.fill = green_fill
            count += 1

        if cell_value != name:
            cell.fill = no_fill

    active_sheet['U1'].value = count
    print(count)
    return count

pitcher_highlight(sheet_name="Texas", column=17, name="Comer")
pitcher_highlight(sheet_name="Milwaukee", column=17, name="Sorensen")

#wb.save(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_AL_Schedule-By_Team.xlsx')
wb.save(r'C:\Users\stoyt\Desktop\1979_AL_Schedule-By_Team.xlsx')

import openpyxl
from openpyxl.styles import PatternFill

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_AL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979 AL Schedule 2.xlsx')

def start_count(sheet_name, pitcher, date):
    current_date_row = 0
    active_sheet = wb[sheet_name]

    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=8, max_col=8):
        cell = row[0]
        if date in str(cell.value) and current_date_row == 0:
            current_date_row = cell.row

    start_count = 0
    total_starts = 0

    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=17, max_col=17):
        cell = row[0]
        cell_value = cell.value

        if pitcher in str(cell_value):
            total_starts += 1

        if cell.row <= current_date_row and pitcher in str(cell_value):
            start_count += 1

    return start_count, total_starts

print(start_count(sheet_name="Chicago", date = "7/5", pitcher="Baumgarten"))
    
#wb.save(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_AL_Schedule-By_Team.xlsx')
wb.save(r'C:\Users\stoyt\Desktop\1979 AL Schedule 2.xlsx')     

import openpyxl

# Assuming 'wb' is a global variable pointing to the workbook
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def start_count(sheet_name, pitcher, date):
    current_date_row = 0
    active_sheet = wb[sheet_name]

    # Find the row corresponding to the current date
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=8, max_col=8):
        cell = row[0]
        if date in str(cell.value) and current_date_row == 0:
            current_date_row = cell.row

    start_count = 0
    total_starts = 0

    # Count starts and total starts
    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=17, max_col=17):
        cell = row[0]
        cell_value = cell.value

        if pitcher in str(cell_value):
            total_starts += 1

        if cell.row <= current_date_row and pitcher in str(cell_value):
            start_count += 1

    return start_count, total_starts

# Example usage
data_sets = [
    {"sheet_name": "San Diego", "date": "6/16", "pitcher": "Perry"},
     {"sheet_name": "Chicago", "date": "6/16", "pitcher": "Holtzman"},
     {"sheet_name": "Cincinnati", "date": "6/16", "pitcher": "Moskau"},
     {"sheet_name": "Philadelphia", "date": "6/16", "pitcher": "Carlton"},
     {"sheet_name": "Houston", "date": "6/16", "pitcher": "Richard"},
     {"sheet_name": "Montreal", "date": "6/16", "pitcher": "Rogers"},
     {"sheet_name": "Pittsburgh", "date": "6/16", "pitcher": "Robinson"},
     {"sheet_name": "Los Angeles", "date": "6/16", "pitcher": "Welch"},
]

active_sheet = wb['More_Info']

row_counter = 1

for data in data_sets:
    pitcher = data["pitcher"]
    date = data["date"]
    team = data["sheet_name"]
    start_count_value, total_starts_value = start_count(team, pitcher, date)

    active_sheet.cell(row=row_counter, column=1).value = date
    active_sheet.cell(row=row_counter, column=2).value = team
    active_sheet.cell(row=row_counter, column=3).value = pitcher
    active_sheet.cell(row=row_counter, column=4).value = start_count_value
    active_sheet.cell(row=row_counter, column=5).value = total_starts_value

    row_counter += 1

# Save the workbook to apply changes
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

# Close the workbook
wb.close()

#print(start_count(sheet_name="Atlanta", date="6/14", pitcher="Solomon"))
#print(start_count(sheet_name="Chicago", date="6/17", pitcher="Krukow"))

import openpyxl
from openpyxl.styles import PatternFill

#wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\CodingDojo\python_stack\algoPractice\1979_AL_Schedule-By_Team.xlsx')
wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')

def find_pitcher(sheet_name, date):
    current_date_row = 0
    active_sheet = wb[sheet_name]

    for row in active_sheet.iter_rows(min_row=3, max_row=165, min_col=8, max_col=8):
        cell = row[0]
        if date in str(cell.value) and current_date_row == 0:
            current_date_row = cell.row
            value_in_column_17 = active_sheet.cell(row=current_date_row, column=17).value
            return value_in_column_17
        
wb.save(r'C:\Users\stoyt\Desktop\1979_NL_Schedule-By_Team.xlsx')