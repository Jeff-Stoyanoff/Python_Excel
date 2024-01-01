import openpyxl
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

def ace_leader_rd1(sheet_name, column):
    active_sheet = wb[sheet_name]
    ace_leader_num = 0
    leader_row = None

    for cell_b in active_sheet[column][1:]:
        cell_value_b = cell_b.value

        if cell_value_b is not None and int(cell_value_b) > ace_leader_num:
            ace_leader_num = int(cell_value_b)
            leader_row = cell_b.row

    if leader_row is not None:
        cell_a = active_sheet.cell(row=leader_row, column=1)  # Access the corresponding cell in Column A
        cell_value_a = cell_a.value

        return cell_value_a, ace_leader_num
    else:
        return None, None

result_a, result_b = ace_leader_rd1('Stats', 'B')

active_sheet = wb['Stats']
active_sheet['Q2'].value = result_a
active_sheet['R2'].value = result_b

print(result_a)
print(result_b)

wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

import openpyxl
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

def set_upsets():
    active_sheet = wb['Results']
    column_AE = active_sheet['AE']
    column_AM = active_sheet['AM']


    for i in range(0, 31, 2):
        A1 = column_AE[i].value
        A2 = column_AE[i+1].value

        print(f'Row {i}: A1 = {A1}, A2 = {A2}')

        if A1 is not None and A2 is not None:

            if A1 > A2:
                column_AM[i].value = "Upset"
            elif A2 > A1:
                column_AM[i+1].value = "Upset"

set_upsets()

wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


wb = load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

def find_player_upset(sheet, start_row, end_row, start_col, end_col, player_col):
    active_sheet = wb[sheet]
    target_player_col = active_sheet[get_column_letter(player_col)]
    upset_winners = []

    for i in range(start_row, end_row, 2):
        cell_value_A = active_sheet.cell(row=i, column=start_col).value
        cell_value_A_next = active_sheet.cell(row=i+1, column=start_col).value
        cell_value_B = active_sheet.cell(row=i, column=end_col).value
        cell_value_B_next = active_sheet.cell(row=i+1, column=end_col).value

        if cell_value_A > cell_value_A_next and cell_value_B == "Upset":
            upset_winner = target_player_col[i-1].value
            upset_winners.append(upset_winner)

        elif cell_value_A_next > cell_value_A and cell_value_B_next == "Upset":
            upset_winner = target_player_col[i].value
            upset_winners.append(upset_winner)

    return upset_winners

print(find_player_upset(sheet='Results', start_row=1, end_row=32, start_col=38, end_col=39, player_col=32))

wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

import openpyxl
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

def update_points_upset(points):
    player_points = []
    upset_winners = find_player_upset(sheet='Results', start_row=1, end_row=32, start_col=38, end_col=39, player_col=32)

    for player in upset_winners:
        player_points.append(points)
    
    combined_list = [item for pair in zip(upset_winners, player_points) for item in pair]

    return combined_list

print(update_points_upset(125))

wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')