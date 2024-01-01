import openpyxl

def create_library(sheet_name_stats, sheet_name_draw):
    ace_leader_library = {}

    # Open the workbook using openpyxl
    wb_openpyxl = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

    # Loop through both "Stats" and "Draw" sheets simultaneously
    active_sheet_stats = wb_openpyxl[sheet_name_stats]
    active_sheet_draw = wb_openpyxl[sheet_name_draw]

    for row_stats, row_draw in zip(active_sheet_stats.iter_rows(min_row=2, max_row=33, min_col=1, max_col=2),
                                  active_sheet_draw.iter_rows(min_row=2, max_row=33, min_col=19, max_col=19)):
        name, aces = row_stats[0].value, row_stats[1].value
        draw_rank = row_draw[0].value

        # Check if the name is already in the library
        if name in ace_leader_library:
            # Update the existing entry
            ace_leader_library[name]['aces'] = aces
            ace_leader_library[name]['rank'] = draw_rank
        else:
            # Add a new entry to the library
            ace_leader_library[name] = {'aces': aces, 'rank': draw_rank}

    # Sort the dictionary based on 'aces' values and create a list of tuples
    sorted_entries = sorted(ace_leader_library.items(), key=lambda x: (x[1]['aces'], x[1]['rank']), reverse=True)

    # Initialize counters for row and column
    column_counter = 1
    row_counter = 37

    # The for loop is now properly indented within the create_library function
    for name, values in sorted_entries:
        if values['aces'] > 7:
            print(f"{name}: {{'Aces': {values['aces']}, 'Rank': {values['rank']}}}" )

            # Access the cell in the current row and column and update its value with Aces
            cell_name = active_sheet_stats.cell(row=row_counter, column=column_counter)
            cell_name.value = name

            # Access the cell in the current row and next column and update its value with Aces
            cell_aces = active_sheet_stats.cell(row=row_counter, column=column_counter + 1)
            cell_aces.value = values['aces']

            # Access the cell in the current row and next column and update its value with Rank
            cell_rank = active_sheet_stats.cell(row=row_counter, column=column_counter + 2)
            cell_rank.value = values['rank']

            # Reset the column counter to 1 and move to the next row
            column_counter = 1
            row_counter += 1

            # Break the loop if you have filled up to row 42
            if row_counter >= 43:
                break

    # Save the workbook to apply changes
    wb_openpyxl.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

    # Close the openpyxl workbook
    wb_openpyxl.close()

# Call the function to execute the code
create_library("Stats", "Draw")




        




