import openpyxl

def create_library(sheet_name_stats, sheet_name_draw): #Returns a list of the highest values for Aces and Breaks along with the name and rank
    ace_leader_library = {}

    # Open the workbook using openpyxl
    wb_openpyxl = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

    # Loop through both "Stats" and "Draw" sheets simultaneously
    active_sheet_stats = wb_openpyxl[sheet_name_stats]
    active_sheet_draw = wb_openpyxl[sheet_name_draw]

    for row_stats, row_draw in zip(active_sheet_stats.iter_rows(min_row=2, max_row=33, min_col=1, max_col=3),
                                   active_sheet_draw.iter_rows(min_row=2, max_row=33, min_col=24, max_col=24)):
        name, aces, breaks = row_stats[0].value, row_stats[1].value, row_stats[2].value
        draw_rank = row_draw[0].value

        # Check if the name is already in the library
        if name in ace_leader_library:
            # Update the existing entry
            ace_leader_library[name]['aces'] = aces
            ace_leader_library[name]['breaks'] = breaks
            ace_leader_library[name]['rank'] = draw_rank
        else:
            # Add a new entry to the library
            ace_leader_library[name] = {'aces': aces, 'breaks': breaks, 'rank': draw_rank}

    # Sort the dictionary based on 'aces' values and create a list of tuples
        sorted_entries = sorted(ace_leader_library.items(), key=lambda x: (x[1]['aces'], x[1]['breaks'], x[1]['rank']), reverse=True)


    # Initialize counters for row and column
    column_counter = 1
    row_counter = 37

    # The for loop is now properly indented within the create_library function
    for name, values in sorted_entries:
        if values['aces'] > 7:
            print(f"{name}: {{'Aces': {values['aces']}, 'Rank': {values['rank']}}}" )

            # Access the cell in the current row and column and update its value with Aces
            cell_name = active_sheet_stats.cell(row=row_counter, column=column_counter)
            cell_name.value = name

            # Access the cell in the current row and next column and update its value with Aces
            cell_aces = active_sheet_stats.cell(row=row_counter, column=column_counter + 1)
            cell_aces.value = values['aces']

            # Access the cell in the current row and next column and update its value with Rank
            cell_rank = active_sheet_stats.cell(row=row_counter, column=column_counter + 2)
            cell_rank.value = values['rank']

            column_counter = 1
            row_counter += 1

            # Break the loop if you have filled up to row 42
            if row_counter >= 44:
                break

    for name, values in sorted_entries:
        if values['breaks'] > 2:
            print(f"{name}: {{'Breaks': {values['breaks']}, 'Rank': {values['rank']}}}" )

            # Access the cell in the current row and column and update its value with Aces
            cell_name = active_sheet_stats.cell(row=row_counter + 9, column=column_counter)
            cell_name.value = name

            # Access the cell in the current row and next column and update its value with Aces
            cell_breaks = active_sheet_stats.cell(row=row_counter + 9, column=column_counter + 1)
            cell_breaks.value = values['breaks']

            # Access the cell in the current row and next column and update its value with Rank
            cell_rank = active_sheet_stats.cell(row=row_counter + 9, column=column_counter + 2)
            cell_rank.value = values['rank']

            # Reset the column counter to 1 and move to the next row
            column_counter = 1
            row_counter += 1

            # Break the loop if you have filled up to row 42
            if row_counter >= 63:
                break

    # Save the workbook to apply changes
    wb_openpyxl.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

    # Close the openpyxl workbook
    wb_openpyxl.close()

# Call the function to execute the code
create_library("Stats", "Draw")

from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

def sort_values(sheet_name, start_row, start_col, end_row, end_col):
    active_sheet = wb[sheet_name]

    rows = []

    for row in active_sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=3):
        row_variable = {'Name': row[0].value, 'Stat': row[1].value, 'Rank': row[2].value}
        rows.append(row_variable)

    sorted_rows = sorted(rows, key=lambda x: (x['Stat'], x['Stat'] + x['Rank']), reverse=True)

    for idx, row in enumerate(sorted_rows, start=start_row):
        active_sheet.cell(row=idx, column=start_col, value=row['Name'])
        active_sheet.cell(row=idx, column=start_col + 1, value=row['Stat'])
        active_sheet.cell(row=idx, column=start_col + 2, value=row['Rank'])

    wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

sort_values('Stats', 37, 5, 43, 7)
sort_values('Stats', 53, 5, 62, 7)

import openpyxl

def update_draw(sheet_one, leader_col, leader_row, leader_max_row, sheet_two, name_col, target_col, starting_points):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

    # Specify the active and update sheets using the provided arguments
    active_sheet = wb[sheet_one]
    update_sheet = wb[sheet_two]

    # Dictionary to store results
    leader_library = {}

    # Set the starting points
    points = starting_points

    # Loop through each cell in the specified column of active_sheet
    for row in active_sheet.iter_rows(min_row=leader_row, max_row=leader_max_row, min_col=leader_col, max_col=leader_col):
        player = row[0].value  # Assuming we're dealing with one cell per row
        if player is not None:
            leader_library[player] = points
        points -= 10  # Subtract 10 for the next player

    # Print the leader_library for verification
    print("Leader Library:")
    for player, points in leader_library.items():
        print(f"{player}: {points}")

    # Loop through the specified range of rows in name_col of update_sheet
    min_row = 141
    max_row = 172
    for row in update_sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=name_col, max_col=name_col + 1):
        name_cell = row[0]  # Assuming we're dealing with one cell per row
        name = name_cell.value
        if name in leader_library:
            # Get the corresponding points from leader_library
            points = leader_library[name]
            # Update the target_col with points for the current name
            target_cell = update_sheet.cell(row=name_cell.row, column=target_col)
            if starting_points > 9:
                target_cell.value = points
            else:
                break

    # Save changes to the workbook
    wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

    # Close the workbook
    wb.close()

# Example usage
update_draw('Stats', 5, 37, 41, 'Draw', 3, 4, 50)
update_draw('Stats', 5, 53, 57, 'Draw', 3, 10, 50)





import openpyxl
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

def stat_leaders_rd1(sheet_name, column):  #Returns the high value from "column" and the corresponding name in col B and col C
    active_sheet = wb[sheet_name]
    leader_num = 0
    leader_row = None

    for cell in active_sheet[column][1:]:
        cell_value = cell.value

        if cell_value is not None and int(cell_value) > leader_num:
            leader_num = int(cell_value)
            leader_row = cell.row

    if leader_row is not None:
        cell_name = active_sheet.cell(row=leader_row, column=1)
        cell_value_name = cell_name.value
        return cell_value_name, leader_num
    else:
        return None, None

result_a, result_b = stat_leaders_rd1('Stats', 'B')
result_c, result_d = stat_leaders_rd1('Stats', 'C')

active_sheet = wb['Stats']
active_sheet['Q2'].value = result_a
active_sheet['R2'].value = result_b
active_sheet['S2'].value = result_c
active_sheet['T2'].value = result_d

wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')


#STUFF WITH UPSETS

import openpyxl
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

def set_upsets():  #Iterates through matchups comparing rankings to establish upsets in a given round
    active_sheet = wb['Raw_Data']
    column_B = active_sheet['B']
    active_sheet_2 = wb['Results']
    column_F = active_sheet_2['AM']

    for i in range(1, 32, 2):
        A1 = column_B[i].value
        A2 = column_B[i+1].value

        print(f'Row {i}: A1 = {A1}, A2 = {A2}')

        if A1 is not None and A2 is not None:
            if A1 > A2:
                column_F[i - 1].value = "Upset"
            elif A2 > A1:
                column_F[i].value = "Upset"

set_upsets()

wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

import openpyxl
from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')

def upset_winners(sheet_name):
    active_sheet = wb[sheet_name]
    
    target_col = 4

    for i in range(1, 33, 2):
        cell_value_A = int(active_sheet.cell(row=i, column=target_col).value)
        cell_value_B = int(active_sheet.cell(row=i+1, column=target_col).value)
        cell_value_rank_top = int(active_sheet.cell(row=i, column=target_col-2).value)
        cell_value_rank_bottom = int(active_sheet.cell(row=i+1, column=target_col-2).value)

        print(cell_value_rank_top)
    
        if cell_value_A > cell_value_B and active_sheet.cell(row=i, column=target_col+2).value == "Upset":
            active_sheet.cell(row=i, column=target_col+3).value = (cell_value_rank_top - cell_value_rank_bottom) + 25
        elif cell_value_B > cell_value_A and active_sheet.cell(row=i+1, column=target_col+2).value == "Upset":
            active_sheet.cell(row=i+1, column=target_col+3).value = (cell_value_rank_bottom - cell_value_rank_top) + 25

upset_winners('Raw_Data')

wb.save(r'C:\Users\stoyt\Desktop\fantasy_tennis_xlsx.xlsx')